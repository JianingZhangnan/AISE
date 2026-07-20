from __future__ import annotations

import difflib
import json
import re
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
from xml.etree import ElementTree

import xlrd
from openpyxl import load_workbook
from pypdf import PdfReader

from phycode.models import (
    PRBENCH_FILE_READ_CHARS,
    FileReadConfig,
    ToolCall,
    ToolResult,
    ToolRiskLevel,
    ToolSpec,
)
from phycode.tools.base import ToolRegistry

MAX_INSPECT_CHARS = 20_000
DEFAULT_FILE_READ_CHARS = PRBENCH_FILE_READ_CHARS
MAX_FILE_READ_CHARS = PRBENCH_FILE_READ_CHARS
MAX_ARCHIVE_MEMBERS = 100
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 50_000_000
MAX_ARCHIVE_MEMBER_BYTES = 10_000_000
MAX_ARCHIVE_MEMBER_CHARS = 8_000
MAX_ARCHIVE_DEPTH = 2
MAX_SPREADSHEET_CELLS = 50_000


def _read(path: Path, limit: int | None, offset: int) -> tuple[str, bool]:
    text = path.read_text(encoding="utf-8")
    sliced = text[offset:]
    if limit is not None and len(sliced) > limit:
        return sliced[:limit], True
    return sliced, False


def _file_read_window(
    args: dict[str, object],
    config: FileReadConfig,
) -> tuple[int, int | None]:
    offset = args.get("offset", 0)
    limit = args.get("limit", config.default_limit)
    if isinstance(offset, bool) or not isinstance(offset, int) or offset < 0:
        raise ValueError("offset must be a non-negative UTF-8 decoded character index")
    if limit is not None:
        if isinstance(limit, bool) or not isinstance(limit, int) or limit < 1:
            raise ValueError("limit must be a positive UTF-8 decoded character count")
        if config.max_limit is not None and limit > config.max_limit:
            raise ValueError(
                f"limit must not exceed {config.max_limit} UTF-8 decoded characters"
            )
    return offset, limit


def _file_read(call: ToolCall, config: FileReadConfig) -> ToolResult:
    offset, limit = _file_read_window(call.args, config)
    content, truncated = _read(Path(call.args["path"]), limit, offset)
    stdout = (
        f"{content}\nnext_offset={offset + len(content)}"
        if truncated and config.emit_next_offset
        else content
    )
    return ToolResult(
        tool_call_id=call.id,
        status="ok",
        stdout=stdout,
        truncated=truncated,
    )


def _file_list(call: ToolCall) -> ToolResult:
    root = Path(call.args.get("path", "."))
    entries = sorted(item.name for item in root.iterdir())
    return ToolResult(tool_call_id=call.id, status="ok", stdout="\n".join(entries))


def _xml_text(xml_bytes: bytes) -> str:
    root = ElementTree.fromstring(xml_bytes)
    return " ".join(text.strip() for text in root.itertext() if text and text.strip())


def _inspect_xlsx_xml(data: bytes) -> str:
    with zipfile.ZipFile(BytesIO(data)) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            shared_strings = [
                " ".join(text.strip() for text in item.itertext() if text and text.strip())
                for item in root
            ]
        lines: list[str] = []
        sheets = sorted(name for name in archive.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        for sheet in sheets:
            root = ElementTree.fromstring(archive.read(sheet))
            lines.append(f"[{sheet}]")
            for row in root.findall(".//{*}row"):
                values: list[str] = []
                for cell in row.findall("{*}c"):
                    coordinate = cell.attrib.get("r", "?")
                    kind = cell.attrib.get("t")
                    value = cell.find("{*}v")
                    inline = cell.find("{*}is")
                    if kind == "s" and value is not None:
                        index = int(value.text or "0")
                        rendered = shared_strings[index] if index < len(shared_strings) else ""
                    elif kind == "inlineStr" and inline is not None:
                        rendered = _xml_text(ElementTree.tostring(inline, encoding="utf-8"))
                    else:
                        rendered = (value.text or "") if value is not None else ""
                    values.append(f"{coordinate}={rendered}")
                if values:
                    lines.append("\t".join(values))
        return "\n".join(lines)


def _spreadsheet_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def _openpyxl_fill(cell: object) -> str | None:
    fill = getattr(cell, "fill", None)
    if fill is None or not fill.fill_type:
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        value = color.rgb
    elif color.type == "indexed" and color.indexed is not None:
        value = f"indexed:{color.indexed}"
    elif color.type == "theme" and color.theme is not None:
        value = f"theme:{color.theme}"
    else:
        value = str(color.type or "unknown")
    return f"{fill.fill_type}:{value}"


def _inspect_xlsx(data: bytes) -> str:
    try:
        formulas = load_workbook(BytesIO(data), data_only=False, read_only=False)
        cached = load_workbook(BytesIO(data), data_only=True, read_only=False)
    except (KeyError, OSError, ValueError, zipfile.BadZipFile):
        return _inspect_xlsx_xml(data)

    lines: list[str] = []
    inspected_cells = 0
    try:
        for sheet in formulas.worksheets:
            lines.append(f"[Sheet: {sheet.title}]")
            cached_sheet = cached[sheet.title]
            for row in sheet.iter_rows():
                row_values: list[str] = []
                for cell in row:
                    inspected_cells += 1
                    if inspected_cells > MAX_SPREADSHEET_CELLS:
                        lines.append(f"[truncated after {MAX_SPREADSHEET_CELLS} cells]")
                        return "\n".join(lines)
                    fill = _openpyxl_fill(cell)
                    cached_value = cached_sheet[cell.coordinate].value
                    value = cached_value if cell.data_type == "f" and cached_value is not None else cell.value
                    if value is None and fill is None:
                        continue
                    rendered = _spreadsheet_value(value) or "<blank>"
                    annotations: list[str] = []
                    if cell.data_type == "f":
                        annotations.append(f"formula={cell.value}")
                    if fill is not None:
                        annotations.append(f"fill={fill}")
                    annotation = f" [{'; '.join(annotations)}]" if annotations else ""
                    row_values.append(f"{cell.coordinate}={rendered}{annotation}")
                if row_values:
                    lines.append("\t".join(row_values))
    finally:
        formulas.close()
        cached.close()
    return "\n".join(lines)


def _inspect_xls(data: bytes) -> str:
    workbook = xlrd.open_workbook(file_contents=data, formatting_info=True)
    lines: list[str] = []
    inspected_cells = 0
    for sheet in workbook.sheets():
        lines.append(f"[Sheet: {sheet.name}]")
        for row_index in range(sheet.nrows):
            row_values: list[str] = []
            for column_index in range(sheet.ncols):
                inspected_cells += 1
                if inspected_cells > MAX_SPREADSHEET_CELLS:
                    lines.append(f"[truncated after {MAX_SPREADSHEET_CELLS} cells]")
                    return "\n".join(lines)
                cell = sheet.cell(row_index, column_index)
                value: object = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(float(cell.value), workbook.datemode)
                fill: str | None = None
                if cell.xf_index is not None and cell.xf_index < len(workbook.xf_list):
                    background = workbook.xf_list[cell.xf_index].background
                    if background is not None and background.fill_pattern:
                        color_index = background.pattern_colour_index
                        rgb = workbook.colour_map.get(color_index)
                        fill = f"pattern:{rgb or color_index}"
                if cell.ctype == xlrd.XL_CELL_EMPTY and fill is None:
                    continue
                coordinate = f"{xlrd.formula.colname(column_index)}{row_index + 1}"
                rendered = _spreadsheet_value(value) or "<blank>"
                annotation = f" [fill={fill}]" if fill is not None else ""
                row_values.append(f"{coordinate}={rendered}{annotation}")
            if row_values:
                lines.append("\t".join(row_values))
    return "\n".join(lines)


def _inspect_office_xml(data: bytes, prefix: str, pattern: str) -> str:
    with zipfile.ZipFile(BytesIO(data)) as archive:
        names = [name for name in archive.namelist() if name.startswith(prefix) and re.search(pattern, name)]
        names.sort(key=lambda name: [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", name)])
        return "\n\n".join(f"[{name}]\n{_xml_text(archive.read(name))}" for name in names)


def _inspect_docx(data: bytes) -> str:
    with zipfile.ZipFile(BytesIO(data)) as archive:
        root = ElementTree.fromstring(archive.read("word/document.xml"))
    body = root.find("{*}body")
    if body is None:
        return _xml_text(ElementTree.tostring(root, encoding="utf-8"))
    lines: list[str] = []
    for block in body:
        if block.tag.endswith("}p"):
            text = " ".join(part.strip() for part in block.itertext() if part and part.strip())
            if text:
                lines.append(text)
        elif block.tag.endswith("}tbl"):
            for row in block.findall("{*}tr"):
                cells = [
                    " ".join(part.strip() for part in cell.itertext() if part and part.strip())
                    for cell in row.findall("{*}tc")
                ]
                if any(cells):
                    lines.append("\t".join(cells))
    return "\n".join(lines)


def _safe_archive_member(name: str) -> bool:
    member = PurePosixPath(name.replace("\\", "/"))
    return not member.is_absolute() and ".." not in member.parts


def _inspect_archive(data: bytes, depth: int) -> str:
    if depth >= MAX_ARCHIVE_DEPTH:
        return "[nested archive depth limit reached]"
    with zipfile.ZipFile(BytesIO(data)) as archive:
        members = [item for item in archive.infolist() if not item.is_dir()]
        if len(members) > MAX_ARCHIVE_MEMBERS:
            raise ValueError(f"Archive has more than {MAX_ARCHIVE_MEMBERS} files")
        total_size = sum(item.file_size for item in members)
        if total_size > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
            raise ValueError(f"Archive expands beyond {MAX_ARCHIVE_UNCOMPRESSED_BYTES} bytes")

        sections: list[str] = []
        for member in members:
            if not _safe_archive_member(member.filename):
                sections.append(f"[Archive member: {member.filename}]\n[skipped unsafe path]")
                continue
            if member.flag_bits & 0x1:
                sections.append(f"[Archive member: {member.filename}]\n[skipped encrypted member]")
                continue
            if member.file_size > MAX_ARCHIVE_MEMBER_BYTES:
                sections.append(f"[Archive member: {member.filename}]\n[skipped oversized member]")
                continue
            try:
                content = _inspect_payload(member.filename, archive.read(member), depth + 1)
            except (ValueError, OSError, zipfile.BadZipFile) as exc:
                content = f"[could not inspect: {exc}]"
            if len(content) > MAX_ARCHIVE_MEMBER_CHARS:
                content = content[:MAX_ARCHIVE_MEMBER_CHARS] + "\n[member output truncated]"
            sections.append(f"[Archive member: {member.filename}]\n{content}")
        return "\n\n".join(sections)


def _inspect_payload(name: str, data: bytes, depth: int = 0) -> str:
    suffix = Path(name).suffix.lower()
    if suffix == ".pdf":
        reader = PdfReader(BytesIO(data))
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)
    if suffix == ".xlsx":
        return _inspect_xlsx(data)
    if suffix == ".xls":
        return _inspect_xls(data)
    if suffix == ".docx":
        return _inspect_docx(data)
    if suffix == ".pptx":
        return _inspect_office_xml(data, "ppt/slides/", r"slide\d+\.xml$")
    if suffix == ".xml":
        try:
            return _xml_text(data)
        except ElementTree.ParseError:
            return data.decode("utf-8", errors="replace")
    if suffix in {
        ".csv",
        ".json",
        ".jsonl",
        ".jsonld",
        ".md",
        ".pdb",
        ".py",
        ".tsv",
        ".txt",
        ".yaml",
        ".yml",
    }:
        text = data.decode("utf-8", errors="replace")
        if suffix in {".json", ".jsonld"}:
            try:
                return json.dumps(json.loads(text), ensure_ascii=False, indent=2)
            except json.JSONDecodeError:
                pass
        return text
    if suffix in {".zip", ".jar", ".whl"}:
        return _inspect_archive(data, depth)
    raise ValueError(f"Unsupported attachment type: {suffix or 'unknown'}; a vision-capable model is required for images")


def _inspect_file(path: Path) -> str:
    return _inspect_payload(path.name, path.read_bytes())


def _file_inspect(call: ToolCall) -> ToolResult:
    content = _inspect_file(Path(call.args["path"]))
    truncated = len(content) > MAX_INSPECT_CHARS
    return ToolResult(
        tool_call_id=call.id,
        status="ok",
        stdout=content[:MAX_INSPECT_CHARS],
        truncated=truncated,
    )


def _file_write(call: ToolCall) -> ToolResult:
    path = Path(call.args["path"])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(str(call.args["content"]), encoding="utf-8")
    return ToolResult(tool_call_id=call.id, status="ok", stdout=f"wrote {path}")


def _file_edit(call: ToolCall) -> ToolResult:
    path = Path(call.args["path"])
    old = str(call.args["old"])
    new = str(call.args["new"])
    before = path.read_text(encoding="utf-8")
    count = before.count(old)
    if count == 0:
        return ToolResult(tool_call_id=call.id, status="tool_error", stderr="old text not found")
    if count > 1:
        return ToolResult(tool_call_id=call.id, status="tool_error", stderr="old text matches more than once")

    after = before.replace(old, new, 1)
    path.write_text(after, encoding="utf-8")
    diff = "\n".join(
        difflib.unified_diff(before.splitlines(), after.splitlines(), fromfile=str(path), tofile=str(path), lineterm="")
    )
    return ToolResult(tool_call_id=call.id, status="ok", stdout=diff)


def register_file_tools(
    registry: ToolRegistry,
    *,
    read_config: FileReadConfig = FileReadConfig(),
) -> None:
    read_description = "Read a file"
    if read_config.emit_next_offset:
        read_description = (
            "Read one bounded text page. offset and limit count zero-based UTF-8 "
            "decoded characters, not line numbers. Follow the returned next_offset "
            "when truncated; do not overlap or repeat previous pages."
        )
    offset_schema: dict[str, object] = {
        "type": "integer",
        "minimum": 0,
        "default": 0,
        "description": "Zero-based UTF-8 decoded character offset.",
    }
    limit_schema: dict[str, object] = {
        "type": "integer",
        "minimum": 1,
        "description": "Maximum UTF-8 decoded characters to return.",
    }
    if read_config.default_limit is not None:
        limit_schema["default"] = read_config.default_limit
    if read_config.max_limit is not None:
        limit_schema["maximum"] = read_config.max_limit
    registry.register(
        ToolSpec(
            name="file.read",
            description=read_description,
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "offset": offset_schema,
                    "limit": limit_schema,
                },
                "required": ["path"],
            },
            risk_level=ToolRiskLevel.SAFE,
        ),
        lambda call: _file_read(call, read_config),
    )
    registry.register(
        ToolSpec(
            name="file.list",
            description="List a directory",
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string"}},
            },
            risk_level=ToolRiskLevel.SAFE,
        ),
        _file_list,
    )
    registry.register(
        ToolSpec(
            name="file.inspect",
            description="Extract text and tabular values from a local PDF, spreadsheet, office document, archive, or text attachment",
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string"}},
                "required": ["path"],
            },
            risk_level=ToolRiskLevel.SAFE,
        ),
        _file_inspect,
    )
    registry.register(
        ToolSpec(
            name="file.write",
            description="Write a file",
            input_schema={
                "type": "object",
                "properties": {"path": {"type": "string"}, "content": {"type": "string"}},
                "required": ["path", "content"],
            },
            risk_level=ToolRiskLevel.RISKY,
            mutates_state=True,
        ),
        _file_write,
    )
    registry.register(
        ToolSpec(
            name="file.edit",
            description="Edit a file by exact replacement",
            input_schema={
                "type": "object",
                "properties": {
                    "path": {"type": "string"},
                    "old": {"type": "string"},
                    "new": {"type": "string"},
                },
                "required": ["path", "old", "new"],
            },
            risk_level=ToolRiskLevel.RISKY,
            mutates_state=True,
        ),
        _file_edit,
    )
