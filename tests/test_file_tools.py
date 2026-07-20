from pathlib import Path
from io import BytesIO
from zipfile import ZipFile

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from phycode.models import PolicyAction, ToolCall
from phycode.policy import PolicyContext
from phycode.tools.base import ToolRegistry, ToolRuntime
from phycode.tools.file_tools import register_file_tools

EXPECTED_FILE_READ_CHARS = 1_200


def test_file_read_reads_workspace_file(tmp_path: Path):
    (tmp_path / "README.md").write_text("hello", encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.read", args={"path": "README.md"}),
        PolicyContext(tmp_path, [], True),
    )
    assert result.policy.decision == PolicyAction.ALLOW
    assert result.tool_result.stdout == "hello"


def test_file_read_resolves_paths_against_policy_workspace(tmp_path: Path):
    (tmp_path / "nested.txt").write_text("workspace copy", encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.read", args={"path": "nested.txt"}),
        PolicyContext(tmp_path, [], True),
    )
    assert result.tool_result.stdout == "workspace copy"


def test_file_read_default_pages_utf8_text_with_recoverable_next_offset(
    tmp_path: Path,
) -> None:
    content = "起点🙂" + "甲乙丙丁" * 700 + "终点🚀"
    (tmp_path / "long.txt").write_text(content, encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)
    runtime = ToolRuntime(registry)
    restored: list[str] = []
    offset = 0
    page_count = 0

    while True:
        page_count += 1
        result = runtime.run(
            ToolCall(
                tool_name="file.read",
                args={"path": "long.txt", "offset": offset},
            ),
            PolicyContext(tmp_path, [], True),
        )
        assert result.tool_result.status == "ok"
        if not result.tool_result.truncated:
            restored.append(result.tool_result.stdout)
            break
        page, separator, marker = result.tool_result.stdout.rpartition("\n")
        assert separator == "\n"
        assert marker.startswith("next_offset=")
        next_offset = int(marker.removeprefix("next_offset="))
        assert 0 < len(page) <= EXPECTED_FILE_READ_CHARS
        assert next_offset == offset + len(page)
        assert len(result.tool_result.stdout) < 1_500
        restored.append(page)
        offset = next_offset

    assert "".join(restored) == content
    assert page_count > 1


def test_file_read_schema_documents_character_paging_contract() -> None:
    registry = ToolRegistry()
    register_file_tools(registry)
    spec = registry.spec_for("file.read")

    assert spec is not None
    description = spec.description.casefold()
    assert "zero-based" in description
    assert "utf-8 decoded characters" in description
    assert "not line" in description
    assert "next_offset" in description
    assert "do not overlap" in description
    offset = spec.input_schema["properties"]["offset"]
    limit = spec.input_schema["properties"]["limit"]
    assert offset["minimum"] == 0
    assert offset["default"] == 0
    assert limit["minimum"] == 1
    assert limit["default"] == EXPECTED_FILE_READ_CHARS
    assert limit["maximum"] == EXPECTED_FILE_READ_CHARS


@pytest.mark.parametrize(
    "args",
    [
        {"offset": True},
        {"offset": -1},
        {"offset": 1.5},
        {"limit": True},
        {"limit": 0},
        {"limit": -1},
        {"limit": EXPECTED_FILE_READ_CHARS + 1},
    ],
)
def test_file_read_rejects_invalid_or_unbounded_windows(
    tmp_path: Path,
    args: dict[str, object],
) -> None:
    (tmp_path / "evidence.txt").write_text("public evidence", encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)

    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.read", args={"path": "evidence.txt", **args}),
        PolicyContext(tmp_path, [], True),
    )

    assert result.tool_result.status in {"invalid_tool_args", "tool_error"}


def test_file_edit_requires_approval_then_writes_diff(tmp_path: Path):
    (tmp_path / "app.py").write_text("x = 1\n", encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)
    call = ToolCall(tool_name="file.edit", args={"path": "app.py", "old": "x = 1", "new": "x = 2"})
    result = ToolRuntime(registry).run(call, PolicyContext(tmp_path, [], True), approved=True)
    assert result.tool_result.status == "ok"
    assert "x = 2" in (tmp_path / "app.py").read_text(encoding="utf-8")
    assert "-x = 1" in result.tool_result.stdout
    assert "+x = 2" in result.tool_result.stdout


def test_file_inspect_extracts_xlsx_values(tmp_path: Path):
    xlsx = tmp_path / "sales.xlsx"
    with ZipFile(xlsx, "w") as archive:
        archive.writestr(
            "xl/worksheets/sheet1.xml",
            "<worksheet xmlns='http://schemas.openxmlformats.org/spreadsheetml/2006/main'>"
            "<sheetData><row r='1'><c r='A1' t='inlineStr'><is><t>Food</t></is></c>"
            "<c r='B1'><v>89706.00</v></c></row></sheetData></worksheet>",
        )
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": "sales.xlsx"}),
        PolicyContext(tmp_path, [], True),
    )

    assert result.policy.decision == PolicyAction.ALLOW
    assert result.tool_result.status == "ok"
    assert "A1=Food\tB1=89706.00" in result.tool_result.stdout


def test_file_inspect_extracts_xlsx_coordinates_formulas_and_fills(tmp_path: Path):
    xlsx = tmp_path / "styled.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet["B2"] = "highlighted"
    sheet["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    sheet["C3"] = "=1+2"
    workbook.save(xlsx)

    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": xlsx.name}),
        PolicyContext(tmp_path, [], True),
    )

    assert "B2=highlighted" in result.tool_result.stdout
    assert "fill=solid:00FFFF00" in result.tool_result.stdout
    assert "C3==1+2 [formula==1+2]" in result.tool_result.stdout


def test_file_inspect_recursively_extracts_supported_archive_members(tmp_path: Path):
    nested = BytesIO()
    with ZipFile(nested, "w") as archive:
        archive.writestr("evidence.txt", "inside nested archive")
    attachment = tmp_path / "bundle.zip"
    with ZipFile(attachment, "w") as archive:
        archive.writestr("facts.xml", "<root>category evidence</root>")
        archive.writestr("nested.zip", nested.getvalue())

    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": attachment.name}),
        PolicyContext(tmp_path, [], True),
    )

    assert "[Archive member: facts.xml]" in result.tool_result.stdout
    assert "category evidence" in result.tool_result.stdout
    assert "[Archive member: evidence.txt]" in result.tool_result.stdout
    assert "inside nested archive" in result.tool_result.stdout


def test_file_inspect_extracts_docx_text(tmp_path: Path):
    docx = tmp_path / "notes.docx"
    with ZipFile(docx, "w") as archive:
        archive.writestr(
            "word/document.xml",
            "<document xmlns='http://schemas.openxmlformats.org/wordprocessingml/2006/main'>"
            "<body><p><r><t>Primary evidence</t></r></p></body></document>",
        )
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": "notes.docx"}),
        PolicyContext(tmp_path, [], True),
    )

    assert result.tool_result.status == "ok"
    assert "Primary evidence" in result.tool_result.stdout


def test_file_inspect_preserves_docx_table_rows(tmp_path: Path):
    docx = tmp_path / "table.docx"
    with ZipFile(docx, "w") as archive:
        archive.writestr(
            "word/document.xml",
            "<document xmlns='http://schemas.openxmlformats.org/wordprocessingml/2006/main'><body><tbl>"
            "<tr><tc><p><r><t>Name</t></r></p></tc><tc><p><r><t>Gift</t></r></p></tc></tr>"
            "<tr><tc><p><r><t>Alex</t></r></p></tc><tc><p><r><t>Book</t></r></p></tc></tr>"
            "</tbl></body></document>",
        )
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": docx.name}),
        PolicyContext(tmp_path, [], True),
    )

    assert "Name\tGift" in result.tool_result.stdout
    assert "Alex\tBook" in result.tool_result.stdout


def test_file_inspect_labels_pptx_slides_in_numeric_order(tmp_path: Path):
    pptx = tmp_path / "slides.pptx"
    with ZipFile(pptx, "w") as archive:
        for number, text in ((10, "Tenth"), (2, "Second"), (1, "First")):
            archive.writestr(
                f"ppt/slides/slide{number}.xml",
                f"<slide xmlns='urn:test'><text>{text}</text></slide>",
            )
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": pptx.name}),
        PolicyContext(tmp_path, [], True),
    )

    output = result.tool_result.stdout
    assert output.index("[ppt/slides/slide1.xml]") < output.index("[ppt/slides/slide2.xml]")
    assert output.index("[ppt/slides/slide2.xml]") < output.index("[ppt/slides/slide10.xml]")


@pytest.mark.parametrize("suffix", [".jsonld", ".pdb"])
def test_file_inspect_reads_gaia_text_attachment_types(tmp_path: Path, suffix: str):
    attachment = tmp_path / f"evidence{suffix}"
    content = '{"name": "GAIA evidence"}' if suffix == ".jsonld" else "HEADER GAIA evidence"
    attachment.write_text(content, encoding="utf-8")
    registry = ToolRegistry()
    register_file_tools(registry)
    result = ToolRuntime(registry).run(
        ToolCall(tool_name="file.inspect", args={"path": attachment.name}),
        PolicyContext(tmp_path, [], True),
    )

    assert result.tool_result.status == "ok"
    assert "GAIA evidence" in result.tool_result.stdout
